# -*- coding: utf-8 -*-
import config
import telebot
import requests
import openpyxl
from openpyxl import load_workbook
from telebot import types
import datetime

global b
global i
global last_num

i = False
b = True

wb = load_workbook('./test.xlsx')

ID = wb['Идентификация пользователя']
CH = wb['Счета на оплату']
RB = wb['Вызов служащих']
PS = wb['Показания счетчиков']

data = {'name':'Name',          #ФИО
        'num':0,                #Номер телефона №1
        'chat_id' : 0,          #ID чата
        'stat' : 'stat',        #Собственник или Сожитель
        'data': 0,              #Дата
        'schet' : 'schet',      #Тип Счётчика
        'kolvo' : 0,            #Количество счётчиков
        'adres': 'None',        #Адрес
        'home': 0,              #Дом
        'numbr': 0,             #Номер телефона №2
        'type':0,               #Тип рабочего
        'pot':0}                #Потребность: Личная или Общая

#$FIO = $_GET['fio'];
#$num = $_GET['num'];
#$stat = $_GET['stat'];

#$message = "";

bot = telebot.TeleBot(config.token)
markup = types.ReplyKeyboardMarkup()
mark = types.ReplyKeyboardMarkup()
markup1 = types.ReplyKeyboardMarkup()
markup2 = types.ReplyKeyboardMarkup()
markup3 = types.ReplyKeyboardMarkup()
markup4 = types.ReplyKeyboardMarkup()
markup5 = types.ReplyKeyboardMarkup()
markup6 = types.ReplyKeyboardMarkup()
markup7 = types.ReplyKeyboardMarkup()
markup2.row("Повторить ввод", "Пройти регистрацию")
markup1.row(config.case1,config.case2)  
markup1.row(config.case3,config.case4)
markup3.row(config.reg1, config.reg2)
markup4.row(config.sch1, config.sch2)
markup4.row(config.sch3, config.sch4)
markup5.row(config.num1, config.num2)
markup5.row(config.num3, config.num4)
markup6.row(config.rab1,config.rab2)
markup6.row(config.rab3)
markup7.row(config.pot1,config.pot2)


def main():

    @bot.message_handler(commands=['start'])    
    def start(message): 
        if(b):
            data['chat_id'] = message.chat.id
            bot.send_message(message.chat.id, config.text)   
            fio()
            
    def fio():
        bot.send_message(data['chat_id'], "Введите ФИО(в формате: Фамилия Имя Отчество)", reply_markup=mark)


    @bot.message_handler(commands=['menu']) 
    def menu(message):
        if(i):
            bot.send_message(data['chat_id'], 'Главное Меню', reply_markup=markup1)

    @bot.message_handler(regexp="\d{9}")
    def prev_reg(message):
        if(i):
            if(edit(message.text)):
                bot.send_message(message.chat.id, "Данные успешно сохранены",reply_markup=markup)

    @bot.message_handler(regexp="\d{11}")    
    def number(message): 
        if(b):
            data['num'] = message.text
            message = data['name']+' ,являясь '+ data['stat'] +' подал заявку на регистрацию. Номер ' + data['num']
            s = requests.get(config.URL + message+'&head='+ config.head1 +'&mail='+config.mail)
            bot.send_message(data['chat_id'], 'Ваша заявка принята. В скором времени вам перезвонит администратор.',reply_markup=markup2)
  
            
    @bot.message_handler(regexp="\d{7}")    
    def number(message): 
        if(b):
            data['num'] = message.text        
            sign_up()

    @bot.message_handler(regexp="\d\d/\d\d/\d{4}")    
    def number(message): 
        if(i):
            data['data'] = message.text
            bot.send_message(message.chat.id, "Введите показание(-я) счётчика(-ов).")


    @bot.message_handler(regexp="\d{1}")
    def prev_reg(message):
        if(b):
            if(message.text == 0):
                reg()
        if(i):
            data['kolvo'] = message.text
            bot.send_message(message.chat.id, "Введите дату проверки счётчика(ДД/ММ/ГГГГ)")


    @bot.message_handler(regexp="\w{2,10}\s\w{2,15}\s\w{2,20}")
    def num1(message):
        if(b):
            data['name'] = message.text                
            bot.send_message(message.chat.id,"Введите номер по договору(если вы хотите зарегистрироваться введите 0)", reply_markup=mark)
        if(i):
            answer(message)


    @bot.message_handler(content_types=['text'])
    def answer(message):
        if(i): 
            
            if(message.text == config.case1):
                print(1)
            if(message.text == config.case2):
                print(cas2())
                pdf = './pdf/' + str(cas2()) + '.pdf'
                f = open(pdf,'rb')
                bot.send_document(data['chat_id'], f,reply_markup=markup)
            if(message.text == config.case3):
                bot.send_message(data['chat_id'], "Выберите тип счётчика", reply_markup=markup4)
            if(message.text == config.case4):
                bot.send_message(data['chat_id'], "Какого сотрудника вы хотите вызвать?", reply_markup=markup6)

            if((message.text == config.sch1)or(message.text == config.sch2)or(message.text == config.sch3)or(message.text == config.sch4)):
                data['schet'] = message.text
                bot.send_message(data['chat_id'], "Сколько счётчиков?", reply_markup=markup5)

            if((message.text == config.rab1)or(message.text == config.rab2)or(message.text == config.rab3)):
                data['type'] = message.text
                bot.send_message(data['chat_id'], "Какая у вас потребность?", reply_markup=markup7)

            if((message.text == config.pot1) or (message.text == config.pot2)):
                data['pot'] = message.text
                if(send()):
                    bot.send_message(data['chat_id'], 'Ваша заявка принята. Мы вам перезвоним для подтверждения вызова.')

        if(b):
            if(message.text == config.reg1) or (message.text == config.reg2):
                data['stat'] = message.text                                       
                bot.send_message(data['chat_id'], "Введите свой телефонный номер в формате 7XXXXXXXXXX.", reply_markup=mark)

            if(message.text == 'Повторить ввод'):
                fio()

            if(message.text == 'Пройти регистрацию'):
                reg()


def send():
    last_num = last2()
    RB['a' + last_num ] = int(last_num) - 1
    RB['b' + last_num ] = data['name']
    RB['c' + last_num ]= data['adres']
    RB['d' + last_num ]= data['home']
    RB['e' + last_num ]= data['numbr']
    RB['f' + last_num ]= data['type']
    RB['g' + last_num ]= data['pot']
    RB['h' + last_num ]= datetime.date.today()
    RB['i' + last_num ]= datetime.time()
    wb.save('test.xlsx')
    message = 'Поступила '+data['pot']+' заявка на '+data['type']+' по адресу: '+data['adres']+', ' + data['home']+'. Номер для подтверждения вызова : '+data['numbr']
    s = requests.get(config.URL + message+'&head='+ config.head2 +'&mail='+config.mail)
    return True
    

def edit(num_text):
    global last_num
    last_num = last()
    PS['a' + last_num ] = int(last_num) - 1
    PS['b' + last_num ] = data['name']
    PS['c' + last_num ]= data['adres']
    PS['d' + last_num ]= data['home']
    PS['e' + last_num ]= data['numbr']
    PS['f' + last_num ]= data['schet']
    PS['i' + last_num ]= int(num_text)
    PS['h' + last_num ]= data['data']
    PS['j' + last_num ]= datetime.date.today()
    PS['i' + last_num ]= time_now()

    wb.save('test.xlsx')
    return True

def time_now():
    return str(datetime.datetime.today().hour) + ':' + str(datetime.datetime.today().minute)


def last():
    for I in range(2,1000):
        if(PS.cell(row = I, column = 1).value == None):
            print(I)
            return str(I)

def last2():
    for I in range(2,1000):
        if(RB.cell(row = I, column = 1).value == None):
            print(I)
            return str(I)

def cas2():
    ans = 0
    for F in range(2, 100):
        if(CH.cell(row = F, column = 2).value == data['name']):
            a = datetime.date.today();m = a.month
            data['adres'] = CH.cell(row = F, column = 3).value
            data['home'] = CH.cell(row = F, column = 4).value
            ans = CH.cell(row = F, column = m+5).value
            break
    return ans   

def error1():
    bot.send_message(data['chat_id'],"Такого номера нет в базе данных дома. Проверьте правильность введённых данных или пройдите регистрацию.",reply_markup=markup2)

def reg():
    bot.send_message(data['chat_id'],"Регистрация.\nВы являетесь собственником или сожителем?",reply_markup=markup3)

def sign_up():
    for I in range(2, 100):
        if((ID.cell(row = I, column = 2).value == data['name']) and (ID.cell(row = I, column = 9).value == eval(data['num']) )):
            global b
            global i
            i = True
            b = False
            data['adres'] = ID.cell(row = I, column = 3).value
            data['home'] = ID.cell(row = I, column = 4).value
            data['numbr'] = ID.cell(row = I, column = 5).value
            markup.row('/menu') 
            bot.send_message(data['chat_id'], 'Добро Пожаловать!!!',reply_markup=markup)
            main()
            return 0
    error1()
            

if __name__ == '__main__':
    main()
    bot.polling(none_stop=True)